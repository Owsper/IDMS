from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, abort, jsonify
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from werkzeug.utils import secure_filename
from email.message import EmailMessage
from database import (
    init_db,
    create_user,
    user_login,
    fetch_unique_email,
    fetch_unique_username,
    get_user_by_id,
    get_user_by_email,
    update_user_profile,
    email_exists_for_other_user,
    username_exists_for_other_user,
    get_dashboard_stats,
    get_recent_activity,
    get_member_statistics,
    get_member_growth_history,
    search_members,
    get_member_filter_options,
    save_upload_metadata,
    get_approved_uploads,
    get_document_categories,
    create_document_category,
    update_document_category,
    get_documents_for_categorization,
    assign_document_category,
    search_approved_documents,
    get_upload_by_id,
    log_document_download,
    mark_user_verified,
    update_user_password,
    get_connection,
    get_table_schema,
    get_import_dashboard_stats,
    create_import_job,
    get_import_job,
    update_import_job,
    replace_import_rows,
    get_import_rows,
    get_import_row,
    update_import_row,
    get_import_history,
)
import uuid
import hashlib
import os
import mimetypes
import csv
import json
import sqlite3
import re
import io
import smtplib
import ssl

try:
    import certifi
except ImportError:  # pragma: no cover - optional dependency
    certifi = None
from datetime import datetime, timedelta
import database

app = Flask(__name__, template_folder="frontend/pages")


def load_local_env(path=".env"):
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


load_local_env()
app.secret_key = os.environ.get("PEXEL_SECRET_KEY", "pexel-dev-secret-key")

# Upload configuration
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB
app.config["UPLOAD_FOLDER"] = os.path.join(os.getcwd(), "secure_uploads")
app.config["PER_FILE_MAX_SIZE"] = 5 * 1024 * 1024  # 5 MB per file
app.config["ALLOWED_EXTENSIONS"] = {".pdf", ".docx", ".doc", ".txt", ".png", ".jpg", ".jpeg"}
app.config["IMPORT_FOLDER"] = os.path.join(app.config["UPLOAD_FOLDER"], "imports")
app.config["IMPORT_ALLOWED_EXTENSIONS"] = {".csv", ".xlsx", ".json", ".sql", ".db", ".sqlite"}
app.config["WHATSAPP_IMPORT_ALLOWED_EXTENSIONS"] = {".txt"}
app.config["IMPORT_MAX_SIZE"] = 10 * 1024 * 1024
app.config["IMPORT_ROLLBACK_MINUTES"] = 60
app.config["IMPORT_BATCH_SIZE"] = 250
app.config["PASSWORD_RESET_MAX_AGE"] = 3600
app.config["DOCUMENT_DOWNLOAD_CACHE_SECONDS"] = 3600
app.config["EMAIL_DELIVERY_MODE"] = os.environ.get("PEXEL_EMAIL_DELIVERY_MODE", "smtp")
app.config["EMAIL_FROM_ADDRESS"] = os.environ.get("PEXEL_EMAIL_FROM_ADDRESS", "")
app.config["EMAIL_FROM_NAME"] = os.environ.get("PEXEL_EMAIL_FROM_NAME", "Pexel")
app.config["SMTP_HOST"] = os.environ.get("PEXEL_SMTP_HOST", "")
app.config["SMTP_PORT"] = int(os.environ.get("PEXEL_SMTP_PORT", "587"))
app.config["SMTP_USERNAME"] = os.environ.get("PEXEL_SMTP_USERNAME", "")
app.config["SMTP_PASSWORD"] = os.environ.get("PEXEL_SMTP_PASSWORD", "")
app.config["SMTP_USE_TLS"] = os.environ.get("PEXEL_SMTP_USE_TLS", "true").lower() in {"1", "true", "yes", "on"}
app.config["SMTP_USE_SSL"] = os.environ.get("PEXEL_SMTP_USE_SSL", "false").lower() in {"1", "true", "yes", "on"}
app.config["EMAIL_TIMEOUT"] = int(os.environ.get("PEXEL_EMAIL_TIMEOUT", os.environ.get("PEXEL_EMAIL_HTTP_TIMEOUT", "10")))

serializer = URLSafeTimedSerializer(app.secret_key)

# Ensure upload directory exists
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["IMPORT_FOLDER"], exist_ok=True)
try:
    os.chmod(app.config["UPLOAD_FOLDER"], 0o700)
except Exception:
    pass

init_db()

# Admin credentials (demo only)
ADMIN_USERNAMES = ["owsper", "christos", "arda", "jira", "salami"]
ADMIN_PASSWORDS = ["owsper", "christos", "arda", "jira", "salami"]


def is_admin_login(email_input, password):
    """Check if credentials match an admin account.
    
    Note: Admin login uses email field but compares against username list.
    This works if admin types their username into the email field.
    """
    if not email_input or not password:
        return False
    username_normalized = email_input.strip().lower()
    for idx, admin_name in enumerate(ADMIN_USERNAMES):
        if username_normalized == admin_name and password == ADMIN_PASSWORDS[idx]:
            return True
    return False


def password_policy_error(password):
    """Return a user-facing policy error, or None when the password is valid."""
    if len(password) < 8:
        return "Password must be at least 8 characters."
    if len(password.encode("utf-8")) > 72:
        return "Password is too long. Please use 72 characters or fewer."
    if not any(character.islower() for character in password):
        return "Password must include a lowercase letter."
    if not any(character.isupper() for character in password):
        return "Password must include an uppercase letter."
    if not any(character.isdigit() for character in password):
        return "Password must include a number."
    return None


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not session.get("user_id") and not session.get("admin_username"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped_view


def admin_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not session.get("admin_username"):
            abort(403)
        return view(*args, **kwargs)
    return wrapped_view


def current_user():
    if session.get("admin_username"):
        admin_name = session["admin_username"]
        return {
            "id": 0,
            "username": admin_name,
            "email": f"{admin_name}@pexel.admin",
            "full_name": f"{admin_name.title()} Admin",
            "bio": "Pexel organizer account.",
            "skills": "Event management, judging, community support",
            "team_role": "Organizer",
            "profile_picture": "",
            "role": "Admin",
            "created_at": "Admin account",
            "updated_at": None,
            "last_login_at": "Current session",
            "is_verified": 1,
        }

    user_id = session.get("user_id")
    return get_user_by_id(user_id) if user_id else None


IMPORT_TARGET_TABLES = {"users_data"}
IMPORT_SYSTEM_COLUMNS = {"created_at", "updated_at", "last_login_at"}


def active_document_category_names():
    return [category["name"] for category in get_document_categories()]


def json_response_error(message, status=400, **extra):
    payload = {"error": message}
    payload.update(extra)
    return jsonify(payload), status


def require_import_job(job_id):
    job = get_import_job(job_id)
    if not job:
        abort(404)
    return job


def safe_target_table(table_name):
    table_name = (table_name or "users_data").strip()
    if table_name not in IMPORT_TARGET_TABLES:
        abort(400, description="Unsupported import target.")
    return table_name


def import_file_path(stored_filename):
    return os.path.join(app.config["IMPORT_FOLDER"], stored_filename)


def parse_csv_file(path):
    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return [{key: value for key, value in row.items()} for row in reader]


def parse_json_file(path):
    with open(path, encoding="utf-8") as handle:
        data = json.load(handle)

    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        rows = next((value for value in data.values() if isinstance(value, list)), [])
    else:
        rows = []

    normalized = []
    for row in rows:
        if isinstance(row, dict):
            normalized.append(row)
    return normalized


def parse_xlsx_file(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("XLSX imports require the openpyxl package.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []

    header_values = [str(value).strip() if value is not None else "" for value in headers]
    records = []
    for values in rows_iter:
        records.append({
            header_values[index]: json_safe_value(value)
            for index, value in enumerate(values)
            if index < len(header_values) and header_values[index]
        })
    return records


def json_safe_value(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def first_user_table(cursor):
    cursor.execute("""
        SELECT name
        FROM sqlite_master
        WHERE type = 'table' AND name NOT LIKE 'sqlite_%'
        ORDER BY name
        LIMIT 1
    """)
    row = cursor.fetchone()
    return row[0] if row else None


def rows_from_sqlite_connection(conn):
    cursor = conn.cursor()
    table_name = first_user_table(cursor)
    if not table_name:
        return []
    cursor.execute(f'SELECT * FROM "{table_name}"')
    columns = [description[0] for description in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def parse_sqlite_file(path):
    conn = sqlite3.connect(path)
    try:
        return rows_from_sqlite_connection(conn)
    finally:
        conn.close()


def parse_sql_file(path):
    with open(path, encoding="utf-8", errors="replace") as handle:
        script = handle.read()
    conn = sqlite3.connect(":memory:")
    try:
        conn.executescript(script)
        return rows_from_sqlite_connection(conn)
    finally:
        conn.close()


def parse_import_file(path, ext):
    if ext == ".csv":
        return parse_csv_file(path)
    if ext == ".json":
        return parse_json_file(path)
    if ext == ".xlsx":
        return parse_xlsx_file(path)
    if ext in {".db", ".sqlite"}:
        return parse_sqlite_file(path)
    if ext == ".sql":
        return parse_sql_file(path)
    raise ValueError("Unsupported import file type.")


def schema_payload(table_name):
    columns = get_table_schema(table_name)
    return [{
        "name": column["name"],
        "type": column["type"],
        "required": bool(column["notnull"]) and column["dflt_value"] is None and not column["pk"],
        "primary_key": bool(column["pk"]),
        "editable": column["name"] not in IMPORT_SYSTEM_COLUMNS,
    } for column in columns]


def guess_field_mapping(headers, schema):
    normalized_headers = {str(header).strip().lower(): header for header in headers}
    mapping = {}
    for column in schema:
        column_name = column["name"]
        if not column["editable"]:
            continue
        source = normalized_headers.get(column_name.lower())
        if source:
            mapping[source] = column_name
    return mapping


def normalize_import_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def validate_type(value, column_type):
    if value is None:
        return True
    column_type = (column_type or "").upper()
    if "INT" in column_type:
        try:
            int(value)
            return True
        except (TypeError, ValueError):
            return False
    if any(token in column_type for token in ["REAL", "FLOA", "DOUB"]):
        try:
            float(value)
            return True
        except (TypeError, ValueError):
            return False
    return True


def coerce_value(value, column_type):
    value = normalize_import_value(value)
    if value is None:
        return None
    column_type = (column_type or "").upper()
    if "INT" in column_type:
        return int(value)
    if any(token in column_type for token in ["REAL", "FLOA", "DOUB"]):
        return float(value)
    return str(value)


def existing_record_for(cursor, table_name, duplicate_key, mapped_data):
    value = mapped_data.get(duplicate_key)
    if value in (None, ""):
        return None
    cursor.execute(f"SELECT * FROM {table_name} WHERE {duplicate_key} = ? LIMIT 1", (value,))
    row = cursor.fetchone()
    return dict(row) if row else None


def validate_import_rows(job, mapping, duplicate_key):
    target_table = safe_target_table(job["target_table"])
    schema = schema_payload(target_table)
    editable_columns = {column["name"]: column for column in schema if column["editable"]}
    rows = parse_import_file(import_file_path(job["stored_filename"]), job["file_ext"])
    prepared = []

    conn = get_connection()
    cursor = conn.cursor()
    try:
        for index, source_row in enumerate(rows, start=1):
            mapped_data = {}
            errors = []
            for source_field, target_field in mapping.items():
                if target_field not in editable_columns:
                    continue
                column = editable_columns[target_field]
                value = normalize_import_value(source_row.get(source_field))
                if not validate_type(value, column["type"]):
                    errors.append(f"{target_field} must match {column['type'] or 'TEXT'}")
                    continue
                try:
                    mapped_data[target_field] = coerce_value(value, column["type"])
                except (TypeError, ValueError):
                    errors.append(f"{target_field} has an invalid value")

            for column in editable_columns.values():
                if column["required"] and mapped_data.get(column["name"]) in (None, ""):
                    errors.append(f"{column['name']} is required")

            email = mapped_data.get("email")
            if email and ("@" not in str(email) or "." not in str(email)):
                errors.append("email must be a valid address")

            if "password" in mapped_data and "password_hash" not in mapped_data:
                errors.append("map passwords to password_hash before import")

            existing = existing_record_for(cursor, target_table, duplicate_key, mapped_data)
            duplicate_value = str(mapped_data.get(duplicate_key) or "")
            status = "invalid" if errors else "valid"

            if not errors and existing:
                comparable_existing = {key: existing.get(key) for key in mapped_data.keys()}
                status = "duplicate" if comparable_existing == mapped_data else "conflict"

            prepared.append({
                "row_number": index,
                "source_data": source_row,
                "mapped_data": mapped_data,
                "status": status,
                "errors": errors,
                "duplicate_key": duplicate_value,
                "existing_record": existing or {},
            })
    finally:
        conn.close()

    return prepared


def row_counts(rows):
    counts = {"total": len(rows), "valid": 0, "invalid": 0, "duplicate": 0, "conflict": 0}
    for row in rows:
        status = row.get("status", "pending")
        counts[status] = counts.get(status, 0) + 1
    return counts


def validate_corrected_import_data(job, values, duplicate_key):
    target_table = safe_target_table(job["target_table"])
    schema = schema_payload(target_table)
    editable_columns = {column["name"]: column for column in schema if column["editable"]}
    unknown_fields = set(values) - set(editable_columns)
    if unknown_fields:
        raise ValueError(f"Unknown import fields: {', '.join(sorted(unknown_fields))}")

    mapped_data = {}
    errors = []
    for field, column in editable_columns.items():
        value = normalize_import_value(values.get(field))
        if not validate_type(value, column["type"]):
            errors.append(f"{field} must match {column['type'] or 'TEXT'}")
            continue
        try:
            mapped_data[field] = coerce_value(value, column["type"])
        except (TypeError, ValueError):
            errors.append(f"{field} has an invalid value")

    for column in editable_columns.values():
        if column["required"] and mapped_data.get(column["name"]) in (None, ""):
            errors.append(f"{column['name']} is required")

    email = mapped_data.get("email")
    if email and ("@" not in str(email) or "." not in str(email)):
        errors.append("email must be a valid address")
    if "password" in mapped_data and "password_hash" not in mapped_data:
        errors.append("map passwords to password_hash before import")

    conn = get_connection()
    cursor = conn.cursor()
    try:
        existing = existing_record_for(cursor, target_table, duplicate_key, mapped_data)
    finally:
        conn.close()

    status = "invalid" if errors else "valid"
    if not errors and existing:
        comparable_existing = {key: existing.get(key) for key in mapped_data}
        status = "duplicate" if comparable_existing == mapped_data else "conflict"
    return mapped_data, errors, status, existing or {}


def generate_verification_token(email):
    return serializer.dumps(email, salt="email-verify")


def confirm_verification_token(token, expiration=3600):
    return serializer.loads(token, salt="email-verify", max_age=expiration)


def account_email_message(purpose, link):
    if purpose == "registration_verification":
        return {
            "subject": "Verify your Pexel account",
            "text": (
                "Welcome to Pexel.\n\n"
                "Use this verification link to verify your account:\n"
                f"{link}\n\n"
                "This link expires in 1 hour."
            ),
            "html": (
                "<p>Welcome to Pexel.</p>"
                "<p>Use this verification link to verify your account:</p>"
                f'<p><a href="{link}">{link}</a></p>'
                "<p>This link expires in 1 hour.</p>"
            ),
        }
    return {
        "subject": "Reset your Pexel password",
        "text": (
            "Use this password reset link to reset your Pexel password:\n"
            f"{link}\n\n"
            "This link expires in 1 hour and can only be used once. "
            "If you did not request this, you can ignore this message."
        ),
        "html": (
            "<p>Use this password reset link to reset your Pexel password:</p>"
            f'<p><a href="{link}">{link}</a></p>'
            "<p>This link expires in 1 hour and can only be used once. "
            "If you did not request this, you can ignore this message.</p>"
        ),
    }


def build_smtp_ssl_context():
    """Build an SSL context, preferring certifi's CA bundle.

    On macOS (and some other environments) the python.org interpreter ships
    without access to the system trust store, so the default context fails to
    verify smtp.gmail.com with CERTIFICATE_VERIFY_FAILED. certifi provides a
    portable CA bundle that fixes this.
    """
    if certifi is not None:
        return ssl.create_default_context(cafile=certifi.where())
    return ssl.create_default_context()


def send_transactional_email(to_email, subject, text_body, html_body):
    mode = app.config.get("EMAIL_DELIVERY_MODE", "smtp")
    if mode == "test":
        return {"sent": True, "provider": "test", "detail": "Test email delivery accepted."}

    if mode != "smtp":
        return {"sent": False, "provider": mode, "detail": "Unsupported email delivery mode."}

    smtp_host = app.config.get("SMTP_HOST", "")
    smtp_port = app.config.get("SMTP_PORT", 587)
    smtp_username = app.config.get("SMTP_USERNAME", "")
    smtp_password = app.config.get("SMTP_PASSWORD", "")
    from_address = app.config.get("EMAIL_FROM_ADDRESS", "")
    if not smtp_host or not from_address:
        return {
            "sent": False,
            "provider": "smtp",
            "detail": (
                "Email delivery is not configured. Set PEXEL_SMTP_HOST and "
                "PEXEL_EMAIL_FROM_ADDRESS."
            ),
        }
    if smtp_username and not smtp_password:
        return {
            "sent": False,
            "provider": "smtp",
            "detail": "Email delivery is not configured. Set PEXEL_SMTP_PASSWORD.",
        }

    from_name = app.config.get("EMAIL_FROM_NAME", "Pexel")
    from_header = f"{from_name} <{from_address}>" if from_name else from_address

    message = EmailMessage()
    message["From"] = from_header
    message["To"] = to_email
    message["Subject"] = subject
    message.set_content(text_body)
    message.add_alternative(html_body, subtype="html")

    try:
        if app.config.get("SMTP_USE_SSL"):
            smtp_context = build_smtp_ssl_context()
            smtp_client = smtplib.SMTP_SSL(
                smtp_host,
                smtp_port,
                timeout=app.config["EMAIL_TIMEOUT"],
                context=smtp_context,
            )
        else:
            smtp_client = smtplib.SMTP(smtp_host, smtp_port, timeout=app.config["EMAIL_TIMEOUT"])

        with smtp_client as server:
            if app.config.get("SMTP_USE_TLS") and not app.config.get("SMTP_USE_SSL"):
                server.starttls(context=build_smtp_ssl_context())
            if smtp_username or smtp_password:
                server.login(smtp_username, smtp_password)
            server.send_message(message)
        return {"sent": True, "provider": "smtp", "detail": "SMTP email delivery accepted."}
    except (OSError, smtplib.SMTPException) as exc:
        return {"sent": False, "provider": "smtp", "detail": str(exc)}


def deliver_account_email(purpose, email, link, user_id=None):
    message = account_email_message(purpose, link)
    delivery = send_transactional_email(
        email,
        message["subject"],
        message["text"],
        message["html"],
    )
    link_id = database.create_auth_email_link(
        purpose,
        email,
        link,
        user_id,
        status="email_sent" if delivery["sent"] else "email_failed",
        error_message=delivery["detail"],
    )
    return {"sent": delivery["sent"], "link_id": link_id, "detail": delivery["detail"]}


def create_verification_email(email, username, user_id=None):
    token = generate_verification_token(email)
    verify_url = url_for("verify_email", token=token, _external=True)
    return deliver_account_email(
        "registration_verification",
        email,
        verify_url,
        user_id=user_id,
    )


def password_reset_fingerprint(user):
    """Tie a reset link to the current password so it can only be used once."""
    return hashlib.sha256(user["password_hash"].encode("utf-8")).hexdigest()[:16]


def generate_password_reset_token(user):
    return serializer.dumps(
        {"user_id": user["id"], "fingerprint": password_reset_fingerprint(user)},
        salt="password-reset",
    )


def confirm_password_reset_token(token):
    payload = serializer.loads(
        token,
        salt="password-reset",
        max_age=app.config["PASSWORD_RESET_MAX_AGE"],
    )
    user = get_user_by_id(payload.get("user_id"))
    if not user or payload.get("fingerprint") != password_reset_fingerprint(user):
        raise BadSignature("Password reset link is no longer valid.")
    return user


def create_password_reset_email(user):
    token = generate_password_reset_token(user)
    reset_url = url_for("reset_password", token=token, _external=True)
    return deliver_account_email(
        "password_reset",
        user["email"],
        reset_url,
        user_id=user["id"],
    )


@app.route("/")
def home():
    return render_template("HomePage.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    error = None
    success = None

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not username or not email or not password or not confirm_password:
            error = "All fields are required."
        elif password != confirm_password:
            error = "Passwords do not match."
        elif password_policy_error(password):
            error = password_policy_error(password)
        else:
            if fetch_unique_username(username):
                error = "Username already exists."
            elif fetch_unique_email(email):
                error = "Email already exists."
            else:
                create_user(username, email, password)
                user = get_user_by_email(email)
                delivery = create_verification_email(email, username, user["id"] if user else None)
                if delivery["sent"]:
                    success = "Account created. Check your email for the verification link."
                else:
                    error = "Account created, but the verification email could not be sent. Ask an administrator to configure email delivery."

    return render_template("RegisterPage.html", error=error, success=success)


@app.route("/verify-email/<token>")
def verify_email(token):
    verification_link = url_for("verify_email", token=token, _external=True)
    try:
        email = confirm_verification_token(token)
    except SignatureExpired:
        return render_template("VerifyPage.html", error="Your verification link has expired.")
    except BadSignature:
        return render_template("VerifyPage.html", error="Invalid verification link.")

    user = get_user_by_email(email)
    if not user:
        return render_template("VerifyPage.html", error="User not found.")

    if int(user.get("is_verified", 0)) == 1:
        return render_template("VerifyPage.html", success="Your account is already verified.")

    verification_email_link = database.get_active_auth_email_link(
        "registration_verification",
        verification_link,
        user_id=user["id"],
    )
    if not verification_email_link:
        return render_template("VerifyPage.html", error="Invalid verification link.")

    mark_user_verified(user["id"])
    database.mark_auth_email_link_used(verification_link)
    return render_template("VerifyPage.html", success="Email verified successfully. You can now log in.")


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")

        if not email or not password:
            error = "All fields are required."
        elif is_admin_login(email, password):
            session.clear()
            session["admin_username"] = email.strip().lower()
            return redirect(url_for("dashboard"))
        else:
            user = user_login(email, password)
            if not user:
                error = "Invalid email or password."
            elif int(user.get("is_verified", 0)) != 1:
                error = "Please verify your email before logging in."
            else:
                session.clear()
                session["user_id"] = user["id"]
                return redirect(url_for("dashboard"))

    return render_template("LoginPage.html", error=error)


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    success = None
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        if email:
            user = get_user_by_email(email)
            if user:
                create_password_reset_email(user)

        # Do not reveal whether an email address is registered.
        success = (
            "If an account exists for that email, a password reset link has been emailed."
        )

    return render_template("PasswordResetPage.html", mode="request", success=success)


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    reset_link = url_for("reset_password", token=token, _external=True)
    try:
        user = confirm_password_reset_token(token)
    except SignatureExpired:
        return render_template(
            "PasswordResetPage.html",
            mode="invalid",
            error="This password reset link has expired. Please request a new one.",
        ), 400
    except BadSignature:
        return render_template(
            "PasswordResetPage.html",
            mode="invalid",
            error="This password reset link is invalid or has already been used.",
        ), 400

    reset_email_link = database.get_active_auth_email_link(
        "password_reset",
        reset_link,
        user_id=user["id"],
    )
    if not reset_email_link:
        return render_template(
            "PasswordResetPage.html",
            mode="invalid",
            error="This password reset link is invalid or has already been used.",
        ), 400

    error = None
    if request.method == "POST":
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        if not password or not confirm_password:
            error = "Both password fields are required."
        elif password != confirm_password:
            error = "Passwords do not match."
        elif password_policy_error(password):
            error = password_policy_error(password)
        else:
            update_user_password(user["id"], password)
            database.mark_auth_email_link_used(reset_link)
            session.clear()
            return render_template(
                "PasswordResetPage.html",
                mode="complete",
                success="Your password has been reset. You can now sign in.",
            )

    return render_template("PasswordResetPage.html", mode="reset", error=error)


@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    stats = get_dashboard_stats(user["id"])
    activities = get_recent_activity(user["id"])
    member_stats = get_member_statistics() if session.get("admin_username") else None
    member_growth = get_member_growth_history() if session.get("admin_username") else []
    import_stats = get_import_dashboard_stats() if session.get("admin_username") else None
    return render_template(
        "DashboardPage.html",
        user=user,
        stats=stats,
        activities=activities,
        member_stats=member_stats,
        member_growth=member_growth,
        import_stats=import_stats,
    )


@app.route("/api/admin/member-stats")
@login_required
@admin_required
def api_admin_member_stats():
    return jsonify(get_member_statistics())


@app.route("/api/admin/member-growth")
@login_required
@admin_required
def api_admin_member_growth():
    return jsonify({"growth": get_member_growth_history()})


@app.route("/admin/document-categories", methods=["GET", "POST"])
@login_required
@admin_required
def admin_document_categories():
    error = None
    success = None
    if request.method == "POST":
        action = request.form.get("action", "").strip()
        try:
            if action in {"create", "update"}:
                name = request.form.get("name", "").strip()
                description = request.form.get("description", "").strip()
                if not 2 <= len(name) <= 50:
                    raise ValueError("Category names must be between 2 and 50 characters.")
                if len(description) > 250:
                    raise ValueError("Category descriptions must be 250 characters or fewer.")
                if action == "create":
                    create_document_category(name, description)
                    success = f"Created category {name}."
                else:
                    category_id = int(request.form.get("category_id", ""))
                    update_document_category(
                        category_id,
                        name,
                        description,
                        request.form.get("is_active") == "1",
                    )
                    success = f"Updated category {name}."
            elif action == "assign":
                assignment = assign_document_category(
                    int(request.form.get("upload_id", "")),
                    int(request.form.get("category_id", "")),
                )
                success = (
                    f"Assigned {assignment['document']['original_filename']} "
                    f"to {assignment['category']['name']}."
                )
            else:
                error = "Unsupported category action."
        except (ValueError, sqlite3.IntegrityError) as exc:
            error = str(exc) if isinstance(exc, ValueError) else "A category with that name already exists."

    all_categories = get_document_categories(include_inactive=True)
    return render_template(
        "AdminDocumentCategoriesPage.html",
        user=current_user(),
        categories=all_categories,
        active_categories=[category for category in all_categories if category["is_active"]],
        documents=get_documents_for_categorization(),
        error=error,
        success=success,
    )


@app.route("/admin/members")
@login_required
@admin_required
def admin_members():
    query = request.args.get("q", "").strip()
    role = request.args.get("role", "").strip()
    team_role = request.args.get("team_role", "").strip()
    verified = request.args.get("verified", "").strip().lower()
    try:
        page = max(1, int(request.args.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    per_page = 25
    results = search_members(
        query=query,
        role=role,
        team_role=team_role,
        verified=verified,
        limit=per_page,
        offset=(page - 1) * per_page,
    )
    return render_template(
        "MembersPage.html",
        user=current_user(),
        members=results["members"],
        total=results["total"],
        query=query,
        role=role,
        team_role=team_role,
        verified=verified,
        page=page,
        per_page=per_page,
        filter_options=get_member_filter_options(),
    )


@app.route("/api/admin/members")
@login_required
@admin_required
def api_admin_members():
    try:
        page = max(1, int(request.args.get("page", 1)))
        per_page = max(1, min(int(request.args.get("per_page", 25)), 100))
    except (TypeError, ValueError):
        return json_response_error("page and per_page must be valid numbers.")

    results = search_members(
        query=request.args.get("q", ""),
        role=request.args.get("role", ""),
        team_role=request.args.get("team_role", ""),
        verified=request.args.get("verified", ""),
        limit=per_page,
        offset=(page - 1) * per_page,
    )
    return jsonify({
        "members": results["members"],
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": results["total"],
            "pages": (results["total"] + per_page - 1) // per_page,
        },
    })


@app.route("/admin/import-data")
@login_required
@admin_required
def admin_import_data():
    user = current_user()
    return render_template(
        "AdminImportDataPage.html",
        user=user,
        import_stats=get_import_dashboard_stats(),
        history=get_import_history(limit=20),
    )


@app.route("/api/admin/import/upload", methods=["POST"])
@login_required
@admin_required
def api_admin_import_upload():
    target_table = safe_target_table(request.form.get("target_table", "users_data"))
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return json_response_error("Choose a database file to import.")

    original_name = secure_filename(uploaded.filename)
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in app.config["IMPORT_ALLOWED_EXTENSIONS"]:
        return json_response_error("Unsupported file type.", allowed=sorted(app.config["IMPORT_ALLOWED_EXTENSIONS"]))

    content = uploaded.read()
    if len(content) > app.config["IMPORT_MAX_SIZE"]:
        return json_response_error("File is too large.", max_size=app.config["IMPORT_MAX_SIZE"])

    stored_name = f"{uuid.uuid4().hex}{ext}"
    dest_path = import_file_path(stored_name)
    with open(dest_path, "wb") as out:
        out.write(content)

    try:
        os.chmod(dest_path, 0o600)
    except Exception:
        pass

    try:
        rows = parse_import_file(dest_path, ext)
    except Exception as exc:
        try:
            os.remove(dest_path)
        except Exception:
            pass
        return json_response_error(str(exc))

    headers = list(rows[0].keys()) if rows else []
    schema = schema_payload(target_table)
    mapping = guess_field_mapping(headers, schema)
    job_id = create_import_job(
        admin_username=session["admin_username"],
        original_filename=original_name,
        stored_filename=stored_name,
        target_table=target_table,
        file_ext=ext,
        file_size=len(content),
    )
    update_import_job(job_id, field_mapping=json.dumps(mapping))

    return jsonify({
        "job_id": job_id,
        "file_name": original_name,
        "row_count": len(rows),
        "headers": headers,
        "schema": schema,
        "suggested_mapping": mapping,
        "duplicate_keys": [column["name"] for column in schema if column["name"] in {"id", "email", "username"}],
        "preview": rows[:20],
    })


@app.route("/api/admin/import/validate", methods=["POST"])
@login_required
@admin_required
def api_admin_import_validate():
    payload = request.get_json(silent=True) or {}
    job = require_import_job(payload.get("job_id"))
    target_table = safe_target_table(job["target_table"])
    schema_names = {column["name"] for column in schema_payload(target_table)}
    duplicate_key = payload.get("duplicate_key") or "email"
    if duplicate_key not in schema_names or duplicate_key not in {"id", "email", "username"}:
        return json_response_error("Duplicate key must be id, email, or username.")

    mapping = payload.get("mapping") or {}
    if not isinstance(mapping, dict) or not mapping:
        return json_response_error("Map at least one uploaded field to the target schema.")

    rows = validate_import_rows(job, mapping, duplicate_key)
    replace_import_rows(job["id"], rows)
    persisted_rows = get_import_rows(job["id"])
    counts = row_counts(rows)
    update_import_job(
        job["id"],
        status="validated",
        field_mapping=json.dumps(mapping),
        duplicate_key=duplicate_key,
        summary=json.dumps(counts),
    )

    return jsonify({
        "job_id": job["id"],
        "counts": counts,
        "invalid_rows": [row for row in persisted_rows if row["status"] == "invalid"][:25],
        "conflicts": [row for row in persisted_rows if row["status"] in {"duplicate", "conflict"}][:25],
        "valid_preview": [row for row in persisted_rows if row["status"] == "valid"][:20],
    })


@app.route("/api/admin/import/<int:job_id>/rows/<int:row_id>", methods=["PATCH"])
@login_required
@admin_required
def api_admin_import_correct_row(job_id, row_id):
    job = require_import_job(job_id)
    row = get_import_row(row_id, job["id"])
    if not row:
        return json_response_error("Import row not found.", status=404)
    if row["status"] != "invalid":
        return json_response_error("Only invalid import rows can be corrected.")

    payload = request.get_json(silent=True) or {}
    mapped_data = payload.get("mapped_data")
    if not isinstance(mapped_data, dict):
        return json_response_error("Corrected mapped_data must be an object.")

    duplicate_key = job.get("duplicate_key") or "email"
    try:
        corrected, errors, status, existing = validate_corrected_import_data(
            job, mapped_data, duplicate_key
        )
    except ValueError as exc:
        return json_response_error(str(exc))

    update_import_row(
        row_id,
        job["id"],
        corrected,
        status,
        errors,
        str(corrected.get(duplicate_key) or ""),
        existing,
    )
    rows = get_import_rows(job["id"])
    counts = row_counts(rows)
    update_import_job(job["id"], status="validated", summary=json.dumps(counts))
    corrected_row = get_import_row(row_id, job["id"])
    return jsonify({
        "row": corrected_row,
        "counts": counts,
        "invalid_rows": [item for item in rows if item["status"] == "invalid"][:25],
        "conflicts": [item for item in rows if item["status"] in {"duplicate", "conflict"}][:25],
        "valid_preview": [item for item in rows if item["status"] == "valid"][:20],
    })


@app.route("/api/admin/import/merge", methods=["POST"])
@login_required
@admin_required
def api_admin_import_merge():
    payload = request.get_json(silent=True) or {}
    job = require_import_job(payload.get("job_id"))
    strategy = payload.get("conflict_strategy", "skip")
    resolutions = payload.get("resolutions") or {}
    if strategy not in {"skip", "overwrite", "manual"}:
        return json_response_error("Conflict strategy must be skip, overwrite, or manual.")

    rows = get_import_rows(job["id"], statuses=["valid", "duplicate", "conflict"])
    if not rows:
        return json_response_error("Validate the import before merging.")

    target_table = safe_target_table(job["target_table"])
    schema = schema_payload(target_table)
    editable_columns = [column["name"] for column in schema if column["editable"]]
    insertable_columns = [name for name in editable_columns if name != "id"]
    duplicate_key = job.get("duplicate_key") or "email"
    now = datetime.utcnow()
    rollback_until = now + timedelta(minutes=app.config["IMPORT_ROLLBACK_MINUTES"])
    summary = {"added": 0, "updated": 0, "skipped": 0, "failed": 0, "manual": 0}

    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("BEGIN")
        for row in rows:
            mapped_data = {
                key: value
                for key, value in row["mapped_data"].items()
                if key in editable_columns and value is not None
            }
            if not mapped_data:
                summary["failed"] += 1
                continue

            existing = existing_record_for(cursor, target_table, duplicate_key, mapped_data)
            decision = strategy
            if strategy == "manual":
                decision = resolutions.get(str(row["id"]), "")
                if decision not in {"skip", "overwrite"}:
                    summary["manual"] += 1
                    continue

            if existing:
                if decision == "skip":
                    summary["skipped"] += 1
                    continue

                update_columns = [key for key in mapped_data.keys() if key in editable_columns and key != "id"]
                if not update_columns:
                    summary["skipped"] += 1
                    continue
                set_clause = ", ".join([f"{column} = ?" for column in update_columns])
                values = [mapped_data[column] for column in update_columns]
                values.append(existing["id"])
                cursor.execute(f"UPDATE {target_table} SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = ?", values)
                cursor.execute("""
                    INSERT INTO import_changes (job_id, table_name, record_pk, action, before_data, after_data)
                    VALUES (?, ?, ?, 'update', ?, ?)
                """, (
                    job["id"],
                    target_table,
                    str(existing["id"]),
                    json.dumps(dict(existing)),
                    json.dumps(mapped_data),
                ))
                summary["updated"] += 1
                continue

            columns = [column for column in insertable_columns if column in mapped_data]
            if "full_name" not in columns and mapped_data.get("username"):
                mapped_data["full_name"] = mapped_data["username"]
                columns.append("full_name")
            if "is_verified" not in columns:
                mapped_data["is_verified"] = 1
                columns.append("is_verified")
            if not columns:
                summary["failed"] += 1
                continue

            placeholders = ", ".join(["?"] * len(columns))
            cursor.execute(
                f"INSERT INTO {target_table} ({', '.join(columns)}) VALUES ({placeholders})",
                [mapped_data[column] for column in columns],
            )
            record_id = cursor.lastrowid
            cursor.execute(f"SELECT * FROM {target_table} WHERE id = ?", (record_id,))
            inserted = dict(cursor.fetchone())
            cursor.execute("""
                INSERT INTO import_changes (job_id, table_name, record_pk, action, before_data, after_data)
                VALUES (?, ?, ?, 'insert', '{}', ?)
            """, (job["id"], target_table, str(record_id), json.dumps(inserted)))
            summary["added"] += 1

        cursor.execute("""
            UPDATE import_jobs
            SET status = ?,
                conflict_strategy = ?,
                summary = ?,
                merged_at = CURRENT_TIMESTAMP,
                rollback_until = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (
            "needs_manual_resolution" if summary["manual"] else "merged",
            strategy,
            json.dumps(summary),
            rollback_until.isoformat(timespec="seconds"),
            job["id"],
        ))
        conn.commit()
    except Exception as exc:
        conn.rollback()
        update_import_job(job["id"], status="failed", error_message=str(exc))
        return json_response_error("Import merge failed and was rolled back.", details=str(exc), status=500)
    finally:
        conn.close()

    return jsonify({"job_id": job["id"], "summary": summary, "rollback_until": rollback_until.isoformat(timespec="seconds")})


@app.route("/api/admin/import/history")
@login_required
@admin_required
def api_admin_import_history():
    return jsonify({"history": get_import_history(limit=50), "stats": get_import_dashboard_stats()})


@app.route("/api/admin/import/<int:job_id>/rollback", methods=["POST"])
@login_required
@admin_required
def api_admin_import_rollback(job_id):
    job = require_import_job(job_id)
    if job.get("rolled_back_at"):
        return json_response_error("This import has already been rolled back.")
    if not job.get("rollback_until"):
        return json_response_error("This import does not have a rollback window.")

    try:
        rollback_until = datetime.fromisoformat(job["rollback_until"])
    except ValueError:
        return json_response_error("Rollback window is invalid.")
    if datetime.utcnow() > rollback_until:
        return json_response_error("Rollback window has expired.")

    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("BEGIN")
        cursor.execute("""
            SELECT *
            FROM import_changes
            WHERE job_id = ? AND rolled_back_at IS NULL
            ORDER BY id DESC
        """, (job_id,))
        changes = [dict(row) for row in cursor.fetchall()]
        for change in changes:
            table_name = safe_target_table(change["table_name"])
            if change["action"] == "insert":
                cursor.execute(f"DELETE FROM {table_name} WHERE id = ?", (change["record_pk"],))
            elif change["action"] == "update":
                before = json.loads(change["before_data"] or "{}")
                columns = [key for key in before.keys() if key not in {"id"}]
                set_clause = ", ".join([f"{column} = ?" for column in columns])
                cursor.execute(
                    f"UPDATE {table_name} SET {set_clause} WHERE id = ?",
                    [before[column] for column in columns] + [change["record_pk"]],
                )
            cursor.execute(
                "UPDATE import_changes SET rolled_back_at = CURRENT_TIMESTAMP WHERE id = ?",
                (change["id"],),
            )

        cursor.execute("""
            UPDATE import_jobs
            SET status = 'rolled_back',
                rolled_back_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (job_id,))
        conn.commit()
    except Exception as exc:
        conn.rollback()
        return json_response_error("Rollback failed.", details=str(exc), status=500)
    finally:
        conn.close()

    return jsonify({"job_id": job_id, "rolled_back": True})


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = current_user()
    error_message = None
    success_message = None

    if session.get("admin_username") and request.method == "POST":
        error_message = "Admin demo profile cannot be edited from this page."
        stats = get_dashboard_stats(user["id"])
        return render_template(
            "ProfilePage.html",
            user=user,
            stats=stats,
            success_message=success_message,
            error_message=error_message,
        )

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        bio = request.form.get("bio", "").strip()
        skills = request.form.get("skills", "").strip()
        team_role = request.form.get("team_role", "").strip()
        profile_picture = request.form.get("profile_picture", "").strip()

        if not full_name or not username or not email:
            error_message = "Full name, username, and email are required."
        elif "@" not in email or "." not in email:
            error_message = "Please enter a valid email address."
        elif username_exists_for_other_user(username, user["id"]):
            error_message = "Username already exists."
        elif email_exists_for_other_user(email, user["id"]):
            error_message = "Email already exists."
        else:
            update_user_profile(
                user["id"],
                full_name,
                username,
                email,
                bio,
                skills,
                team_role or "Developer",
                profile_picture,
            )
            success_message = "Profile updated successfully."
            user = current_user()

    stats = get_dashboard_stats(user["id"])
    return render_template(
        "ProfilePage.html",
        user=user,
        stats=stats,
        success_message=success_message,
        error_message=error_message,
    )


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/import-files", methods=["GET", "POST"])
@login_required
def import_files():
    user = current_user()
    categories = active_document_category_names()

    def render_page(**context):
        stats = get_dashboard_stats(user["id"]) if user else {}
        return render_template(
            "ImportFilesPage.html",
            user=user,
            stats=stats,
            categories=categories,
            **context,
        )

    if request.method == "POST":
        category = request.form.get("category", "General").strip()
        if category not in categories:
            return render_page(error="Select a valid document category.")
        if "files" not in request.files:
            return render_page(error="No files provided.")

        files = request.files.getlist("files")
        saved_files = []
        is_admin = bool(session.get("admin_username"))

        for f in files:
            if not f or f.filename == "":
                continue

            original_name = secure_filename(f.filename)
            ext = os.path.splitext(original_name)[1].lower()

            if ext not in app.config["ALLOWED_EXTENSIONS"]:
                return render_page(error=f"Invalid file type: {ext}")

            content = f.read()
            if len(content) > app.config["PER_FILE_MAX_SIZE"]:
                return render_page(error=f"File too large: {original_name}")

            sha256 = hashlib.sha256(content).hexdigest()
            stored_name = f"{uuid.uuid4().hex}{ext}"
            dest_path = os.path.join(app.config["UPLOAD_FOLDER"], stored_name)

            try:
                with open(dest_path, "wb") as out:
                    out.write(content)
                os.chmod(dest_path, 0o600)
            except Exception:
                return render_page(error="Failed to save file on server.")

            try:
                save_upload_metadata(
                    user_id=(None if is_admin else (user["id"] if user else None)),
                    original_filename=original_name,
                    stored_filename=stored_name,
                    mime_type=f.mimetype or "application/octet-stream",
                    size=len(content),
                    sha256=sha256,
                    approved=(1 if is_admin else 0),
                    approved_by=(session.get("admin_username") if is_admin else None),
                    category=category,
                )
            except Exception:
                app.logger.exception("Failed saving upload metadata")
                try:
                    os.remove(dest_path)
                except Exception:
                    pass
                return render_page(error="Failed to save upload metadata.")

            saved_files.append(original_name)

        if is_admin:
            return render_page(success=f"Uploaded and approved {len(saved_files)} files.")
        else:
            return render_page(success=f"Uploaded {len(saved_files)} files; pending approval.")

    return render_page()


@app.route("/files")
@login_required
def list_files():
    user = current_user()
    files = get_approved_uploads(limit=200)
    categories = active_document_category_names()
    return render_template(
        "DocumentsPage.html",
        user=user,
        files=files,
        categories=categories,
        stats=get_dashboard_stats(user["id"]),
    )


@app.route("/api/documents")
@login_required
def api_documents():
    query = request.args.get("q", "").strip()
    requested_category = request.args.get("category", "").strip()
    categories = active_document_category_names()
    category = next(
        (name for name in categories if name.casefold() == requested_category.casefold()),
        "",
    )
    if len(query) > 150:
        return json_response_error("Search query must be 150 characters or fewer.")
    if requested_category and not category:
        return json_response_error("Select a valid document category.")
    try:
        page = max(1, int(request.args.get("page", 1)))
        per_page = max(1, min(int(request.args.get("per_page", 25)), 100))
    except (TypeError, ValueError):
        return json_response_error("page and per_page must be valid numbers.")

    results = search_approved_documents(
        query=query,
        category=category,
        limit=per_page,
        offset=(page - 1) * per_page,
    )
    documents = []
    for document in results["documents"]:
        title = document.pop("original_filename")
        documents.append({
            **document,
            "title": title,
            "download_url": url_for("download_file", file_id=document["id"]),
        })

    return jsonify({
        "documents": documents,
        "query": query,
        "category": category,
        "categories": categories,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": results["total"],
            "pages": (results["total"] + per_page - 1) // per_page,
        },
    })


@app.route("/files/<int:file_id>/download")
@login_required
def download_file(file_id):
    record = get_upload_by_id(file_id)
    if not record or int(record.get("approved", 0)) != 1:
        abort(404)

    stored_name = record["stored_filename"]
    response = send_from_directory(
        app.config["UPLOAD_FOLDER"],
        stored_name,
        as_attachment=True,
        mimetype=record.get("mime_type"),
        download_name=record.get("original_filename") or stored_name,
        conditional=True,
        etag=True,
        max_age=app.config["DOCUMENT_DOWNLOAD_CACHE_SECONDS"],
    )
    response.cache_control.public = False
    response.cache_control.private = True
    response.vary.add("Cookie")

    if request.method == "GET" and response.status_code in {200, 206}:
        log_document_download(
            upload_id=record["id"],
            user_id=session.get("user_id"),
            admin_username=session.get("admin_username"),
            ip_address=request.remote_addr or "",
            user_agent=str(request.user_agent)[:500],
        )
    return response


def parse_form_datetime(value):
    if not value:
        raise ValueError("Date and time are required.")
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise ValueError("Date and time must be valid.") from exc


def parse_optional_date(value, field_name):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).date().isoformat()
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a valid date.") from exc


def current_actor_name():
    user = current_user() or {}
    return user.get("username") or session.get("admin_username") or "system"


WHATSAPP_TIMESTAMP_FORMATS = (
    "%m/%d/%y %I:%M %p",
    "%m/%d/%Y %I:%M %p",
    "%d/%m/%y %I:%M %p",
    "%d/%m/%Y %I:%M %p",
    "%d/%m/%y %H:%M",
    "%d/%m/%Y %H:%M",
    "%m/%d/%y %H:%M",
    "%m/%d/%Y %H:%M",
)
WHATSAPP_LINE_PATTERNS = (
    re.compile(r"^\[(?P<date>\d{1,2}/\d{1,2}/\d{2,4}), (?P<time>\d{1,2}:\d{2}(?::\d{2})?(?:\s?[AP]M)?)\] (?P<sender>[^:]+): (?P<message>.*)$"),
    re.compile(r"^(?P<date>\d{1,2}/\d{1,2}/\d{2,4}), (?P<time>\d{1,2}:\d{2}(?::\d{2})?(?:\s?[AP]M)?) - (?P<sender>[^:]+): (?P<message>.*)$"),
)
WHATSAPP_SYSTEM_MARKERS = (
    "messages and calls are end-to-end encrypted",
    "this message was deleted",
    "you deleted this message",
)


def parse_whatsapp_datetime(date_value, time_value):
    normalized_time = time_value.upper().replace("\u202f", " ").replace(".", "").strip()
    normalized_time = re.sub(r"\s+", " ", normalized_time)
    for fmt in WHATSAPP_TIMESTAMP_FORMATS:
        try:
            return datetime.strptime(f"{date_value} {normalized_time}", fmt)
        except ValueError:
            pass
    return None


def clean_whatsapp_text(value):
    value = (value or "").replace("\ufeff", "").replace("\u200e", "").replace("\u200f", "")
    lines = [re.sub(r"\s+", " ", line).strip() for line in value.splitlines()]
    return "\n".join(line for line in lines if line).strip()


def classify_whatsapp_message(message):
    lowered = message.lower()
    if "image omitted" in lowered or "photo omitted" in lowered:
        return "image"
    if "video omitted" in lowered:
        return "video"
    if "audio omitted" in lowered or "voice message omitted" in lowered:
        return "audio"
    if "sticker omitted" in lowered:
        return "sticker"
    if "document omitted" in lowered or "file omitted" in lowered:
        return "document"
    if "contact card omitted" in lowered or "contact omitted" in lowered:
        return "contact"
    if "location:" in lowered:
        return "location"
    if "omitted" in lowered or "<media" in lowered:
        return "media"
    return "text"


def parse_whatsapp_export(text, filename=""):
    messages = []
    current = None

    def append_current():
        if not current:
            return
        message = clean_whatsapp_text(current["message"])
        sender = clean_whatsapp_text(current["sender"])
        lowered = message.lower()
        if not sender or not message:
            return
        if any(marker in lowered for marker in WHATSAPP_SYSTEM_MARKERS):
            return
        messages.append({
            "sender": sender,
            "sent_at": current["sent_at"].isoformat(timespec="seconds"),
            "message": message,
            "media_type": classify_whatsapp_message(message),
            "source_filename": filename,
        })

    for raw_line in text.splitlines():
        line = raw_line.strip("\ufeff").rstrip()
        if not line:
            continue
        match = next((pattern.match(line) for pattern in WHATSAPP_LINE_PATTERNS if pattern.match(line)), None)
        if not match:
            if any(marker in line.lower() for marker in WHATSAPP_SYSTEM_MARKERS):
                continue
            if current:
                current["message"] = f"{current['message']}\n{line.strip()}"
            continue

        append_current()
        sender = match.group("sender").strip()
        message = match.group("message").strip()
        parsed_at = parse_whatsapp_datetime(match.group("date"), match.group("time"))
        if not parsed_at:
            current = None
            continue
        current = {
            "sender": sender,
            "sent_at": parsed_at,
            "message": message,
        }

    append_current()
    return messages


def parse_whatsapp_upload(uploaded):
    if not uploaded or not uploaded.filename:
        raise ValueError("Choose a WhatsApp .txt export.")

    original_name = secure_filename(uploaded.filename)
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in app.config["WHATSAPP_IMPORT_ALLOWED_EXTENSIONS"]:
        raise ValueError("WhatsApp imports must be .txt files.")

    content = uploaded.read()
    if len(content) > app.config["IMPORT_MAX_SIZE"]:
        raise ValueError("File is too large.")

    messages = parse_whatsapp_export(content.decode("utf-8", errors="replace"), original_name)
    if not messages:
        raise ValueError("No WhatsApp messages were found in this export.")
    return messages


def csv_download(filename, rows, fieldnames):
    handle = io.StringIO()
    writer = csv.DictWriter(handle, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow({field: row.get(field, "") for field in fieldnames})
    return app.response_class(
        handle.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/voting", methods=["GET", "POST"])
@login_required
def voting():
    error = None
    success = None
    if request.method == "POST":
        try:
            if request.form.get("action") == "create":
                if not session.get("admin_username"):
                    abort(403)
                options = request.form.get("options", "").replace("\r", "").split("\n")
                database.create_voting_event(
                    request.form.get("title", ""),
                    request.form.get("description", ""),
                    options,
                    parse_form_datetime(request.form.get("start_at", "")),
                    parse_form_datetime(request.form.get("end_at", "")),
                    created_by=current_actor_name(),
                    eligibility={
                        "membership_status": request.form.get("membership_status", "verified"),
                        "min_membership_days": request.form.get("min_membership_days", 0),
                        "allowed_roles": request.form.get("allowed_roles", ""),
                    },
                )
                success = "Voting event created."
            elif request.form.get("action") == "vote":
                database.cast_vote(
                    int(request.form.get("event_id", "")),
                    int(request.form.get("option_id", "")),
                    session.get("user_id"),
                    app.secret_key,
                )
                success = "Vote recorded securely."
        except (TypeError, ValueError) as exc:
            error = str(exc)
    return render_template(
        "VotingPage.html",
        user=current_user(),
        events=database.list_voting_events(user_id=session.get("user_id")),
        is_admin=bool(session.get("admin_username")),
        error=error,
        success=success,
    )


@app.route("/api/voting/events", methods=["POST"])
@login_required
@admin_required
def api_voting_create_event():
    payload = request.get_json(silent=True) or {}
    try:
        event_id = database.create_voting_event(
            payload.get("title", ""),
            payload.get("description", ""),
            payload.get("options", []),
            parse_form_datetime(payload.get("start_at", "")),
            parse_form_datetime(payload.get("end_at", "")),
            created_by=current_actor_name(),
            eligibility=payload.get("eligibility") or {},
        )
    except ValueError as exc:
        return json_response_error(str(exc))
    return jsonify({"event_id": event_id}), 201


@app.route("/api/voting/votes", methods=["POST"])
@login_required
def api_voting_cast_vote():
    payload = request.get_json(silent=True) or {}
    try:
        database.cast_vote(int(payload.get("event_id")), int(payload.get("option_id")), session.get("user_id"), app.secret_key)
    except (TypeError, ValueError) as exc:
        return json_response_error(str(exc))
    return jsonify({"status": "recorded"})


@app.route("/api/voting/events/<int:event_id>/results")
@login_required
@admin_required
def api_voting_results(event_id):
    try:
        results = database.get_voting_results(event_id)
    except ValueError as exc:
        return json_response_error(str(exc), status=404)
    return jsonify(results)


@app.route("/api/voting/events/<int:event_id>/results.csv")
@login_required
@admin_required
def api_voting_results_csv(event_id):
    try:
        results = database.get_voting_results(event_id)
    except ValueError as exc:
        return json_response_error(str(exc), status=404)
    event = results["event"]
    rows = []
    for option in results["options"]:
        rows.append({
            "event_id": event["id"],
            "event_title": event["title"],
            "option_id": option["id"],
            "option_label": option["label"],
            "votes": option["votes"],
            "percentage": option["percentage"],
            "winner": "yes" if option["winner"] else "no",
            "total_votes": results["total_votes"],
            "is_tie": "yes" if results["is_tie"] else "no",
        })
    return csv_download(
        "voting-results.csv",
        rows,
        [
            "event_id",
            "event_title",
            "option_id",
            "option_label",
            "votes",
            "percentage",
            "winner",
            "total_votes",
            "is_tie",
        ],
    )


@app.route("/whatsapp-analytics", methods=["GET", "POST"])
@login_required
@admin_required
def whatsapp_analytics_page():
    error = None
    success = None
    if request.method == "POST":
        try:
            messages = parse_whatsapp_upload(request.files.get("file"))
            database.store_whatsapp_messages(messages)
            success = f"Imported {len(messages)} messages."
        except (TypeError, ValueError) as exc:
            error = str(exc)
    return render_template("WhatsAppAnalyticsPage.html", user=current_user(), analytics=database.whatsapp_analytics(), error=error, success=success)


@app.route("/api/whatsapp/import", methods=["POST"])
@login_required
@admin_required
def api_whatsapp_import():
    try:
        messages = parse_whatsapp_upload(request.files.get("file"))
    except ValueError as exc:
        return json_response_error(str(exc))
    return jsonify({"imported": database.store_whatsapp_messages(messages)})


@app.route("/api/whatsapp/analytics")
@login_required
@admin_required
def api_whatsapp_analytics():
    return jsonify(database.whatsapp_analytics())


@app.route("/notifications", methods=["GET", "POST"])
@login_required
def notifications_page():
    if request.method == "POST":
        if not session.get("admin_username"):
            abort(403)
        database.create_notification(
            request.form.get("category", "general"),
            request.form.get("title", ""),
            request.form.get("body", ""),
            channel=request.form.get("channel", "in-app"),
            scheduled_for=request.form.get("scheduled_for") or None,
        )
    conn = get_connection()
    rows = [dict(row) for row in conn.execute("SELECT * FROM notifications ORDER BY created_at DESC LIMIT 100").fetchall()]
    conn.close()
    return render_template("NotificationsPage.html", user=current_user(), notifications=rows, is_admin=bool(session.get("admin_username")))


@app.route("/meetings", methods=["GET", "POST"])
@login_required
def meetings():
    error = None
    success = None
    if request.method == "POST":
        try:
            action = request.form.get("action")
            if action == "create":
                if not session.get("admin_username"):
                    abort(403)
                database.create_meeting(
                    request.form.get("title", ""),
                    request.form.get("description", ""),
                    parse_form_datetime(request.form.get("meeting_at", "")),
                    request.form.get("location", ""),
                    request.form.get("agenda", ""),
                    [item.strip() for item in request.form.get("invitees", "").split(",") if item.strip()],
                    created_by=current_actor_name(),
                    meeting_type=request.form.get("meeting_type", "general"),
                )
                success = "Meeting scheduled."
            elif action == "attendance":
                if not session.get("admin_username"):
                    abort(403)
                database.record_attendance(int(request.form.get("meeting_id")), int(request.form.get("member_id")), request.form.get("status", "present"))
                success = "Attendance recorded."
            elif action == "minutes":
                if not session.get("admin_username"):
                    abort(403)
                database.add_meeting_minutes(int(request.form.get("meeting_id")), request.form.get("title", ""), request.form.get("content", ""), uploaded_by=current_actor_name())
                success = "Minutes saved."
        except (TypeError, ValueError) as exc:
            error = str(exc)
    members = search_members(limit=100, offset=0)["members"] if session.get("admin_username") else []
    return render_template(
        "MeetingsPage.html",
        user=current_user(),
        meetings=database.list_meetings(),
        members=members,
        attendance=database.meeting_attendance_summary(),
        is_admin=bool(session.get("admin_username")),
        min_meeting_at=datetime.utcnow().strftime("%Y-%m-%dT%H:%M"),
        error=error,
        success=success,
    )


@app.route("/api/meetings")
@login_required
def api_meetings():
    try:
        start = parse_optional_date(request.args.get("start"), "start")
        end = parse_optional_date(request.args.get("end"), "end")
        if start and end and start > end:
            return json_response_error("start date must be before end date.")
    except ValueError as exc:
        return json_response_error(str(exc))
    return jsonify({"meetings": database.list_meetings(start, end)})


@app.route("/api/meetings", methods=["POST"])
@login_required
@admin_required
def api_meeting_create():
    payload = request.get_json(silent=True) or {}
    try:
        meeting_id = database.create_meeting(
            payload.get("title", ""),
            payload.get("description", ""),
            parse_form_datetime(payload.get("meeting_at", "")),
            payload.get("location", ""),
            payload.get("agenda", ""),
            payload.get("invitees", []),
            current_actor_name(),
            payload.get("meeting_type", "general"),
        )
    except (TypeError, ValueError) as exc:
        return json_response_error(str(exc))
    return jsonify({"meeting_id": meeting_id, "meeting": database.get_meeting(meeting_id)}), 201


@app.route("/api/meetings/attendance.csv")
@login_required
@admin_required
def api_meeting_attendance_csv():
    return csv_download("attendance-report.csv", database.meeting_attendance_summary(), ["label", "present", "total"])


@app.route("/financial", methods=["GET", "POST"])
@login_required
@admin_required
def financial():
    error = None
    success = None
    if request.method == "POST":
        try:
            if request.form.get("action") == "transaction":
                database.create_transaction(request.form.get("transaction_date", ""), request.form.get("type", ""), request.form.get("category", ""), request.form.get("amount", 0), request.form.get("description", ""), current_actor_name())
                success = "Transaction recorded."
            elif request.form.get("action") == "budget":
                database.upsert_budget(request.form.get("category", ""), request.form.get("allocated_amount", 0), request.form.get("fiscal_period", ""))
                success = "Budget saved."
        except ValueError as exc:
            error = str(exc)
    return render_template("FinancialPage.html", user=current_user(), report=database.financial_report(), error=error, success=success)


@app.route("/api/financial/report")
@login_required
@admin_required
def api_financial_report():
    return jsonify(database.financial_report())


@app.route("/api/financial/report.csv")
@login_required
@admin_required
def api_financial_report_csv():
    report = database.financial_report()
    return csv_download("financial-report.csv", report["monthly"], ["label", "income", "expense"])


@app.route("/activity-summary")
@login_required
@admin_required
def activity_summary_page():
    period = request.args.get("period", "monthly")
    return render_template("ActivitySummaryPage.html", user=current_user(), summary=database.activity_summary(period))


@app.route("/api/admin/activity-summary")
@login_required
@admin_required
def api_activity_summary():
    return jsonify(database.activity_summary(request.args.get("period", "monthly")))


@app.route("/bugs", methods=["GET", "POST"])
@login_required
def bugs_page():
    error = None
    if request.method == "POST":
        try:
            if request.form.get("action") == "create":
                database.create_bug_report(request.form.get("title", ""), request.form.get("severity", "Medium"), request.form.get("steps", ""), request.form.get("expected", ""), request.form.get("actual", ""), current_actor_name())
            elif request.form.get("action") == "status":
                if not session.get("admin_username"):
                    abort(403)
                database.update_bug_status(int(request.form.get("bug_id")), request.form.get("status", "Open"), request.form.get("resolution_notes", ""))
        except ValueError as exc:
            error = str(exc)
    return render_template("BugTrackerPage.html", user=current_user(), bugs=database.list_bug_reports(), is_admin=bool(session.get("admin_username")), error=error)


@app.route("/help")
@login_required
def help_page():
    return render_template("HelpPage.html", user=current_user())


@app.route("/developer-guide")
@login_required
@admin_required
def developer_guide_page():
    return render_template("DeveloperGuidePage.html", user=current_user())


if __name__ == "__main__":
    app.run(debug=True)
