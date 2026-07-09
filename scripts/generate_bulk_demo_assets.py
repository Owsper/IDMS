import csv
import html
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "generated"
OUT.mkdir(exist_ok=True)

MEMBER_XLSX = OUT / "idms_4000_existing_members.xlsx"
MEMBER_CSV = OUT / "idms_4000_existing_members.csv"
ORG_HTML = OUT / "idms_organization_documents_1000_pages.html"
ORG_MD = OUT / "idms_organization_documents_1000_pages.md"

TEMP_PASSWORD = "TempPass123!"
TEMP_PASSWORD_HASH = "$2b$12$JYLoBFE9kieENkvbKFnvAe1.pAs51o5O2qyMsFQxZQF6LeHXNB/3m"

FIRST_NAMES = [
    "Ava", "Mateo", "Mina", "Noah", "Zoe", "Liam", "Sara", "Ethan",
    "Amara", "Jonas", "Priya", "Theo", "Nora", "Leo", "Iris", "Omar",
    "Elena", "Kai", "Maya", "Rafael", "Sofia", "Daniel", "Lena", "Arjun",
]

LAST_NAMES = [
    "Chen", "Rivera", "Patel", "Kim", "Martin", "Okafor", "Nguyen", "Brooks",
    "Singh", "Muller", "Garcia", "Hassan", "Brown", "Kowalski", "Silva", "Taylor",
    "Ivanov", "Rossi", "Novak", "Mensah", "Costa", "Ahmed", "Wilson", "Tan",
]

TEAM_ROLES = [
    "Developer", "Designer", "QA Engineer", "Coordinator", "Analyst",
    "Operations", "Finance", "Documentation", "Support", "Team Lead",
]

SKILL_SETS = [
    "Python, Flask, SQLite",
    "HTML, CSS, JavaScript",
    "Testing, Playwright, Bug Reports",
    "UX, Figma, Research",
    "SQL, Spreadsheets, Analytics",
    "Operations, Scheduling, Support",
    "Documentation, Training, Review",
    "Finance, Budgets, Reporting",
    "Security, Code Review, APIs",
    "Community, Outreach, Moderation",
]

DOC_TYPES = [
    "Governance Charter",
    "Membership Policy",
    "Document Control Procedure",
    "Admin Import Procedure",
    "Security Guideline",
    "Meeting Operations Manual",
    "Finance Control Policy",
    "Voting Procedure",
    "Bug Triage Standard",
    "Activity Reporting Guide",
]

MODULES = [
    "Authentication and User Management",
    "Document Library",
    "Admin Data Import",
    "Voting",
    "Meetings",
    "Finance",
    "Notifications",
    "WhatsApp Analytics",
    "Activity Summary",
    "Bug Tracking",
    "Teams and Team Chat",
]


def member_row(index):
    first = FIRST_NAMES[index % len(FIRST_NAMES)]
    last = LAST_NAMES[(index * 7) % len(LAST_NAMES)]
    full_name = f"{first} {last}"
    username = f"{first.lower()}-{last.lower()}-{index:04d}"
    email = f"{first.lower()}.{last.lower()}.{index:04d}@example.org"
    role = "Admin" if index % 97 == 0 else "Participant"
    verified = 0 if index % 29 == 0 else 1
    opt_in = 0 if index % 17 == 0 else 1
    team_role = TEAM_ROLES[index % len(TEAM_ROLES)]
    skills = SKILL_SETS[index % len(SKILL_SETS)]
    created_at = datetime(2025, 1, 1) + timedelta(days=index % 540, hours=index % 24)
    updated_at = created_at + timedelta(days=index % 45)
    last_login_at = "" if index % 13 == 0 else (updated_at + timedelta(hours=2)).isoformat(sep=" ", timespec="seconds")
    return {
        "id": 100000 + index,
        "username": username,
        "email": email,
        "password_hash": TEMP_PASSWORD_HASH,
        "full_name": full_name,
        "bio": f"Synthetic IDMS member record for {team_role.lower()} workflow testing.",
        "skills": skills,
        "team_role": team_role,
        "profile_picture": "",
        "role": role,
        "is_verified": verified,
        "notification_opt_in": opt_in,
        "created_at": created_at.isoformat(sep=" ", timespec="seconds"),
        "updated_at": updated_at.isoformat(sep=" ", timespec="seconds"),
        "last_login_at": last_login_at,
    }


def generate_members():
    headers = [
        "id",
        "username",
        "email",
        "password_hash",
        "full_name",
        "bio",
        "skills",
        "team_role",
        "profile_picture",
        "role",
        "is_verified",
        "notification_opt_in",
        "created_at",
        "updated_at",
        "last_login_at",
    ]
    rows = [member_row(i) for i in range(1, 4001)]

    with MEMBER_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Existing Members"
    sheet.append(headers)
    for row in rows:
        sheet.append([row[header] for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = {
        "A": 10,
        "B": 26,
        "C": 34,
        "D": 64,
        "E": 24,
        "F": 58,
        "G": 34,
        "H": 20,
        "J": 16,
        "M": 22,
        "N": 22,
        "O": 22,
    }
    for column in range(1, len(headers) + 1):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = widths.get(letter, 14)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    meta = workbook.create_sheet("Import Notes")
    meta.append(["Field", "Value"])
    meta.append(["Target table", "users_data"])
    meta.append(["Rows", len(rows)])
    meta.append(["Temporary password for all generated members", TEMP_PASSWORD])
    meta.append(["Duplicate key suggestion", "email"])
    meta.append(["Data status", "Synthetic example data, not real people"])
    for cell in meta[1]:
        cell.font = Font(bold=True)
    meta.column_dimensions["A"].width = 42
    meta.column_dimensions["B"].width = 80

    workbook.save(MEMBER_XLSX)


def page_body(page):
    doc_type = DOC_TYPES[(page - 1) % len(DOC_TYPES)]
    module = MODULES[(page * 3) % len(MODULES)]
    owner = TEAM_ROLES[(page * 5) % len(TEAM_ROLES)]
    revision = f"ORG-{page:04d}"
    controls = [
        f"Purpose: define repeatable organization practice for {module.lower()} inside IDMS.",
        "Scope: applies to administrators, verified members, project maintainers, and support roles.",
        "Storage: approved documents are uploaded through the IDMS document library and categorized for retrieval.",
        "Access: member-facing material is available after login; restricted administrative material requires admin access.",
        "Evidence: activity logs, import history, download logs, meeting minutes, and bug reports provide traceability.",
    ]
    checklist = [
        "Confirm responsible owner and backup owner.",
        "Review current route, database helper, and template behavior before changing the process.",
        "Record decisions in meeting minutes or the activity summary.",
        "Update user-facing guidance when workflow screens or validations change.",
        "Retest the affected module with a representative member account and an admin account.",
    ]
    risks = [
        "Unverified accounts should not bypass protected member workflows.",
        "Bulk imports must be validated before merge and rollback metadata must be preserved.",
        "Downloaded documents should remain auditable through download records.",
        "Financial changes require positive values, clear categories, and budget review.",
        "Voting events require clear eligibility rules and one-vote enforcement.",
    ]
    return {
        "title": f"{doc_type}: {module}",
        "revision": revision,
        "owner": owner,
        "controls": controls,
        "checklist": checklist,
        "risks": risks,
    }


def generate_documents():
    generated_at = datetime.now().isoformat(sep=" ", timespec="seconds")
    styles = """
    <style>
      @page { size: A4; margin: 18mm; }
      body { font-family: Arial, Helvetica, sans-serif; color: #17202a; margin: 0; background: #f4f6f8; }
      .page { box-sizing: border-box; min-height: 297mm; padding: 18mm; page-break-after: always; background: #fff; }
      .kicker { color: #536471; font-size: 11px; text-transform: uppercase; letter-spacing: .08em; }
      h1 { font-size: 26px; margin: 8px 0 12px; }
      h2 { font-size: 15px; margin: 18px 0 8px; border-bottom: 1px solid #d8dee4; padding-bottom: 4px; }
      p, li { font-size: 12px; line-height: 1.55; }
      .meta { display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px; margin: 12px 0 18px; }
      .meta div { border: 1px solid #d8dee4; padding: 8px; font-size: 11px; }
      .label { display: block; color: #536471; font-size: 10px; text-transform: uppercase; }
      footer { margin-top: 18px; color: #536471; font-size: 10px; }
    </style>
    """
    html_parts = [
        "<!doctype html>",
        "<html>",
        "<head>",
        '<meta charset="utf-8">',
        "<title>IDMS Organization Documents - 1000 Pages</title>",
        styles,
        "</head>",
        "<body>",
    ]
    md_parts = [
        "# IDMS Organization Documents - 1000 Pages",
        "",
        f"Generated: {generated_at}",
        "",
        "This synthetic document pack is for IDMS testing, demos, and document-library population.",
        "",
    ]

    for page in range(1, 1001):
        body = page_body(page)
        html_parts.extend([
            '<section class="page">',
            f'<div class="kicker">IDMS Organization Document Pack / Page {page:04d} of 1000</div>',
            f"<h1>{html.escape(body['title'])}</h1>",
            '<div class="meta">',
            f"<div><span class=\"label\">Document ID</span>{body['revision']}</div>",
            f"<div><span class=\"label\">Owner</span>{html.escape(body['owner'])}</div>",
            "<div><span class=\"label\">Status</span>Approved for demo use</div>",
            f"<div><span class=\"label\">Generated</span>{html.escape(generated_at)}</div>",
            "</div>",
            "<p>IDMS is a Flask and SQLite based organization management system for members, documents, teams, voting, meetings, finance, notifications, activity reporting, WhatsApp analytics, imports, and bug tracking. This page defines a synthetic but realistic operating document for project demonstrations and document-management testing.</p>",
            "<h2>Operating Controls</h2>",
            "<ul>",
        ])
        html_parts.extend(f"<li>{html.escape(item)}</li>" for item in body["controls"])
        html_parts.extend(["</ul>", "<h2>Workflow Checklist</h2>", "<ol>"])
        html_parts.extend(f"<li>{html.escape(item)}</li>" for item in body["checklist"])
        html_parts.extend(["</ol>", "<h2>Risks And Review Notes</h2>", "<ul>"])
        html_parts.extend(f"<li>{html.escape(item)}</li>" for item in body["risks"])
        html_parts.extend([
            "</ul>",
            "<h2>Approval Note</h2>",
            "<p>This generated page may be uploaded to the document module as sample organization content. Replace it with real governance text before production use.</p>",
            f"<footer>IDMS synthetic organization document pack / {body['revision']} / Page {page:04d}</footer>",
            "</section>",
        ])

        md_parts.extend([
            f"<!-- PAGE {page:04d} -->",
            "",
            f"## Page {page:04d}: {body['title']}",
            "",
            f"- Document ID: {body['revision']}",
            f"- Owner: {body['owner']}",
            "- Status: Approved for demo use",
            "",
            "### Operating Controls",
            "",
        ])
        md_parts.extend(f"- {item}" for item in body["controls"])
        md_parts.extend(["", "### Workflow Checklist", ""])
        md_parts.extend(f"{idx}. {item}" for idx, item in enumerate(body["checklist"], start=1))
        md_parts.extend(["", "### Risks And Review Notes", ""])
        md_parts.extend(f"- {item}" for item in body["risks"])
        md_parts.extend(["", ""])

    html_parts.extend(["</body>", "</html>"])
    ORG_HTML.write_text("\n".join(html_parts), encoding="utf-8")
    ORG_MD.write_text("\n".join(md_parts), encoding="utf-8")


def main():
    generate_members()
    generate_documents()
    print(f"Created {MEMBER_XLSX}")
    print(f"Created {MEMBER_CSV}")
    print(f"Created {ORG_HTML}")
    print(f"Created {ORG_MD}")
    print(f"Temporary password for generated members: {TEMP_PASSWORD}")


if __name__ == "__main__":
    main()
